from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOODER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOODER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'Minha_Panilha'

worksheet: Worksheet = workbook[sheet_name]

for row in worksheet.iter_rows():
    for cell in row:
        print(cell)
    
# workbook.save(WORKBOOK_PATH)
