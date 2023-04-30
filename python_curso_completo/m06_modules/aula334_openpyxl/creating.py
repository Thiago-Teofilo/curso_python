from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOODER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOODER / 'workbook.xlsx'

workbook = Workbook()

sheet_name = 'Minha_Panilha'
workbook.create_sheet(sheet_name, 0)
worksheet: Worksheet = workbook[sheet_name]
# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10]
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_col in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_col)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)